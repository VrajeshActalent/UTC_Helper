"""
backend.py — Core processing logic.

Parses a .c test file and fills 'Expected Results' (J:L) and
'Actual Results' (M:O) columns in an Excel workbook.

Two macro syntaxes supported
─────────────────────────────
  1. EXPECTED_CALLS("func#inst;func#inst;...")
       • Single string argument
       • Multiple functions separated by semicolons
       • Instance encoded after # in the function token
       • Tokens may carry a numeric prefix:  33*FuncName#1

  2. EXPECT_CALL("FuncName", "qualifier", "instance")
       • Three separate string arguments
       • One call per line
       • Function name in first arg (may carry N* prefix: "33*FuncName")
       • Second arg (qualifier) is ignored
       • Instance in third arg

  ✘  EXPECTED_CALL / EXPECT_CALLS — intentionally ignored

Deduplication
─────────────
  Duplicate (func, instance) pairs within the same TC block are removed
  (first occurrence kept, order preserved).

Return-value extraction  (3 scenarios)
───────────────────────────────────────
  Scenario 1 – /* Stub for function X */
  Scenario 2 – /* Replace-Wrapper for function X */ / REPLACE_X()
  Scenario 3 – void stub → no retval clause in sentence

Numbering
─────────
  Every function line is ALWAYS numbered (1. 2. 3. …), even single items.
  Append: existing content is retroactively numbered if needed;
          new lines continue from max+1.

NA handling
───────────
  Cells containing only "NA" (any case) are treated as empty → overwritten.
"""

from __future__ import annotations
import re
from copy import copy
from openpyxl import load_workbooks
from openpyxl.styles import Alignment

x =20
# ─────────────────────────────────────────────────────────────────────────────
# Regex patterns for the two supported macro syntaxes
# ─────────────────────────────────────────────────────────────────────────────

# Syntax 1: EXPECTED_CALLS("func#inst;func#inst;...")
_EXPECTED_CALLS_RE = re.compile(
    r'\bEXPECTED_CALLS\s*\(\s*"([^"]*)"\s*\)',
    re.DOTALL
)

# Syntax 2: EXPECT_CALL("FuncName", "qualifier", "instance")
_EXPECT_CALL_RE = re.compile(
    r'\bEXPECT_CALL\s*\(\s*"([^"]*)"\s*,\s*"[^"]*"\s*,\s*"([^"]*)"\s*\)',
    re.DOTALL
)

# ─────────────────────────────────────────────────────────────────────────────
# Token helpers
# ─────────────────────────────────────────────────────────────────────────────

def _strip_numeric_prefix(name: str) -> str:
    """
    Strip leading  N*  from a function name or token.
    Works for both syntaxes:
      '33*FuncName#1' → 'FuncName#1'   (EXPECTED_CALLS token)
      '33*FuncName'   → 'FuncName'     (EXPECT_CALL first arg)
    """
    return re.sub(r'^\d+\*', '', name.strip())


def _deduplicate(funcs: list) -> list:
    """Remove duplicate (func, instance) pairs, preserving first-occurrence order."""
    seen, result = set(), []
    for item in funcs:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Return-value extraction  (stub / wrapper parser)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_body(text: str, brace_pos: int) -> str:
    """Return content between the outermost { } starting at brace_pos."""
    depth, body_start = 0, -1
    for i in range(brace_pos, len(text)):
        if text[i] == '{':
            depth += 1
            if depth == 1:
                body_start = i + 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                return text[body_start:i]
    return text[body_start:] if body_start != -1 else ""


def _extract_instance_returnvalues(body: str) -> dict:
    """
    Scan a stub/wrapper body.
    Returns { instance_str -> return_value_str | None }.
    None = void stub (no returnValue assignment in that IF_INSTANCE block).
    """
    inst_map: dict[str, str | None] = {}
    if_re = re.compile(r'IF_INSTANCE\(\s*"(\d+)"\s*\)\s*\{', re.DOTALL)
    for im in if_re.finditer(body):
        inst      = im.group(1)
        brace_pos = body.index('{', im.start())
        block     = _extract_body(body, brace_pos)
        m = re.search(r'\breturnValue\s*=\s*([^;]+);', block)
        inst_map[inst] = m.group(1).strip() if m else None
    return inst_map


def _parse_stub_blocks(content: str) -> dict:
    """
    Build { func_name -> { instance -> return_value | None } }
    from Stub and Replace-Wrapper definitions in the file.
    """
    rv_map: dict[str, dict] = {}

    block_re = re.compile(
        r'/\*\s*(?P<kind>Stub|Before-Wrapper|After-Wrapper|Replace-Wrapper)'
        r'\s+for\s+function\s+(?P<fname>\w+)\s*\*/',
        re.DOTALL
    )
    for bm in block_re.finditer(content):
        kind, fname = bm.group('kind'), bm.group('fname')
        brace_pos   = content.find('{', bm.end())
        if brace_pos == -1:
            continue
        body = _extract_body(content, brace_pos)
        if kind in ('Stub', 'Replace-Wrapper'):
            rv_map.setdefault(fname, {}).update(_extract_instance_returnvalues(body))

    # REPLACE_ functions without a preceding comment
    replace_re = re.compile(r'\bREPLACE_(\w+)\s*\([^)]*\)\s*\{')
    for rm in replace_re.finditer(content):
        fname     = rm.group(1)
        brace_pos = content.index('{', rm.start())
        body      = _extract_body(content, brace_pos)
        rv_map.setdefault(fname, {}).update(_extract_instance_returnvalues(body))

    return rv_map


def get_return_value(rv_map: dict, func_name: str, instance: str) -> str | None:
    """Return the return-value string for func_name@instance, or None."""
    return rv_map.get(func_name, {}).get(instance)


# ─────────────────────────────────────────────────────────────────────────────
# C file parser  (test-case blocks)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_tc_block(block_text: str) -> list[tuple[str, str]]:
    """
    Extract all (func_name, instance) pairs from one test-case block.
    Handles both macro syntaxes and strips N* prefixes.
    Returns a deduplicated list.
    """
    funcs: list[tuple[str, str]] = []

    # ── Syntax 1: EXPECTED_CALLS("func#inst;...") ─────────────────────────
    for ec_m in _EXPECTED_CALLS_RE.finditer(block_text):
        calls_str = ec_m.group(1).strip()
        if not calls_str:
            continue
        for raw_token in calls_str.split(";"):
            token = _strip_numeric_prefix(raw_token)   # strip N* from token
            if not token:
                continue
            if "#" in token:
                func, inst = token.rsplit("#", 1)
                funcs.append((func.strip(), inst.strip()))
            else:
                funcs.append((token.strip(), "1"))

    # ── Syntax 2: EXPECT_CALL("FuncName", "qualifier", "instance") ────────
    for ec_m in _EXPECT_CALL_RE.finditer(block_text):
        raw_func = ec_m.group(1)          # may carry N* prefix
        inst     = ec_m.group(2).strip()
        func     = _strip_numeric_prefix(raw_func)   # strip N* from func name
        if func:
            funcs.append((func, inst))

    return _deduplicate(funcs)


def parse_c_file(c_path: str) -> tuple[dict, dict]:
    """
    Parse the .c file.

    Returns
    -------
    tc_map : { tc_id -> [(func_name, instance), ...] }
    rv_map : { func_name -> { instance -> return_value | None } }
    """
    with open(c_path, "r", encoding="utf-8") as fh:
        content = fh.read()

    block_re = re.compile(r'START_TEST\(\s*"([^"]+)"', re.DOTALL)
    tc_map: dict[str, list] = {}
    starts = list(block_re.finditer(content))

    for idx, m in enumerate(starts):
        tc_id      = m.group(1).strip().split(":")[0].strip()
        block_end  = starts[idx + 1].start() if idx + 1 < len(starts) else len(content)
        block_text = content[m.start():block_end]
        tc_map[tc_id] = _parse_tc_block(block_text)

    rv_map = _parse_stub_blocks(content)
    return tc_map, rv_map


# ─────────────────────────────────────────────────────────────────────────────
# Serial-number helpers
# ─────────────────────────────────────────────────────────────────────────────

def get_next_serial(text: str) -> int:
    """
    Return the next serial number to continue from existing cell content.
      empty / NA          → 1
      numbered lines      → max(N) + 1
      un-numbered content → 2  (existing treated as implicit item 1)
    """
    if not text or not text.strip():
        return 1
    matches = re.findall(r"^\s*(\d+)\.", text, re.MULTILINE)
    return max(int(x) for x in matches) + 1 if matches else 2


def prefix_unnumbered(text: str) -> str:
    """Prepend '1. ' to un-numbered existing content so numbering is consistent."""
    if not text or not text.strip():
        return text
    if re.search(r"^\s*\d+\.", text, re.MULTILINE):
        return text          # already numbered — leave as-is
    return "1. " + text.strip()


# ─────────────────────────────────────────────────────────────────────────────
# Sentence builders  — ALWAYS numbered
# ─────────────────────────────────────────────────────────────────────────────

def _expected_sentence(func: str, inst: str, retval: str | None) -> str:
    if retval is not None:
        return (f'Function "{func}" should be called with instance "{inst}"'
                f' with returnvalue "{retval}".')
    return f'Function "{func}" should be called with instance "{inst}".'


def _actual_sentence(func: str, inst: str, retval: str | None) -> str:
    if retval is not None:
        return (f'Function "{func}" is called with instance "{inst}"'
                f' with returnvalue "{retval}".')
    return f'Function "{func}" is called with instance "{inst}".'


def build_expected_lines(funcs: list, rv_map: dict, start_serial: int) -> str:
    """Build Expected Results lines. Always numbered: '1. Function ...'"""
    lines = []
    for i, (func, inst) in enumerate(funcs):
        retval = get_return_value(rv_map, func, inst)
        lines.append(f"{start_serial + i}. " + _expected_sentence(func, inst, retval))
    return "\n".join(lines)


def build_actual_lines(funcs: list, rv_map: dict, start_serial: int) -> str:
    """Build Actual Results lines. Always numbered: '1. Function ...'"""
    lines = []
    for i, (func, inst) in enumerate(funcs):
        retval = get_return_value(rv_map, func, inst)
        lines.append(f"{start_serial + i}. " + _actual_sentence(func, inst, retval))
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

def col_index(letter: str) -> int:
    idx = 0
    for ch in letter.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx


def find_header_row(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == "TC No.":
                return cell.row
    return None


def find_tc_col(sheet, header_row: int):
    for cell in sheet[header_row]:
        if cell.value and str(cell.value).strip() == "TC No.":
            return cell.column
    return None


def ensure_merge(sheet, row: int, c1: int, c2: int):
    for mr in sheet.merged_cells.ranges:
        if (mr.min_row == row and mr.max_row == row
                and mr.min_col == c1 and mr.max_col == c2):
            return
    sheet.merge_cells(start_row=row, start_column=c1,
                      end_row=row,   end_column=c2)


def get_cell_text(sheet, row: int, col: int) -> str:
    """
    Read existing text. Returns "" for blank or 'NA' (any case) cells
    so they are overwritten rather than appended to.
    """
    v = sheet.cell(row=row, column=col).value
    if not v:
        return ""
    text = str(v).strip()
    return "" if text.upper() == "NA" else text


def write_cell(sheet, row: int, c1: int, c2: int, text: str, sample=None):
    ensure_merge(sheet, row, c1, c2)
    cell = sheet.cell(row=row, column=c1)
    cell.value = text
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    if sample and sample.font:
        cell.font = copy(sample.font)


# ─────────────────────────────────────────────────────────────────────────────
# Main processing entry point
# ─────────────────────────────────────────────────────────────────────────────

def run_processing(xlsx_path: str, c_path: str, log) -> int:
    """
    Fill / append Expected Results and Actual Results in the Excel workbook.

    Parameters
    ----------
    xlsx_path : str            Path to the Excel workbook (edited in-place).
    c_path    : str            Path to the C test file.
    log       : callable(str)  Progress-message callback.

    Returns
    -------
    int  Number of rows updated.
    """
    EXP_S, EXP_E = col_index("J"), col_index("L")
    ACT_S, ACT_E = col_index("M"), col_index("O")

    # ── Parse C file ──────────────────────────────────────────────────────────
    log("Parsing C file…")
    tc_map, rv_map = parse_c_file(c_path)
    log(f"  Found {len(tc_map)} test case(s).")
    log(f"  Found return-value data for {len(rv_map)} stub/wrapper function(s).")
    for fname, inst_vals in rv_map.items():
        log(f"    {fname}: {inst_vals}")

    # ── Load workbook ─────────────────────────────────────────────────────────
    log("\nLoading Excel workbook…")
    wb    = load_workbook(xlsx_path)
    sheet = wb.active

    header_row = find_header_row(sheet)
    if header_row is None:
        raise ValueError("Could not find 'TC No.' header in the Excel sheet.")

    tc_col = find_tc_col(sheet, header_row)
    if tc_col is None:
        raise ValueError("Could not find 'TC No.' column in the Excel sheet.")

    log(f"  Header at row {header_row}, TC No. at column {tc_col}.\n")

    sample_cell = sheet.cell(row=header_row + 1, column=EXP_S)
    matched     = 0

    # ── Iterate data rows ─────────────────────────────────────────────────────
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        tc_value = sheet.cell(row=row_idx, column=tc_col).value
        if not tc_value:
            continue

        tc_str = str(tc_value).strip()

        # Resolve TC key  e.g. "877_MC:PDV_Init" → "877_MC"
        m = re.match(r"(\d+_MC)", tc_str.replace(" ", ""))
        tc_key = m.group(1) if m else tc_str

        if tc_key not in tc_map:
            num = re.match(r"(\d+)", tc_str)
            if num:
                tc_key = next(
                    (k for k in tc_map if k.startswith(num.group(1) + "_")),
                    None
                )

        if not tc_key or tc_key not in tc_map:
            log(f"  Row {row_idx}: '{tc_str}' → no match in C file")
            continue

        funcs = tc_map[tc_key]
        if not funcs:
            log(f"  Row {row_idx}: '{tc_str}' → matched but no EXPECTED_CALLS / EXPECT_CALL entries found")
            continue

        # ── Expected Results ──────────────────────────────────────────────────
        ex_exp     = get_cell_text(sheet, row_idx, EXP_S)
        exp_serial = get_next_serial(ex_exp)

        if ex_exp:
            base_exp  = prefix_unnumbered(ex_exp)
            final_exp = base_exp + "\n" + build_expected_lines(funcs, rv_map, exp_serial)
        else:
            final_exp = build_expected_lines(funcs, rv_map, 1)

        # ── Actual Results ────────────────────────────────────────────────────
        ex_act     = get_cell_text(sheet, row_idx, ACT_S)
        act_serial = get_next_serial(ex_act)

        if ex_act:
            base_act  = prefix_unnumbered(ex_act)
            final_act = base_act + "\n" + build_actual_lines(funcs, rv_map, act_serial)
        else:
            final_act = build_actual_lines(funcs, rv_map, 1)

        write_cell(sheet, row_idx, EXP_S, EXP_E, final_exp, sample_cell)
        write_cell(sheet, row_idx, ACT_S, ACT_E, final_act, sample_cell)

        total_lines = max(final_exp.count("\n"), final_act.count("\n")) + 1
        sheet.row_dimensions[row_idx].height = max(30, total_lines * 30)

        action = "Appended" if ex_exp else "Filled"
        log(f"  Row {row_idx}: '{tc_str}' → {action} "
            f"({len(funcs)} function(s), serial from {exp_serial})")
        matched += 1

    wb.save(xlsx_path)
    log(f"\n✅  Done — updated {matched} row(s). Saved to:\n    {xlsx_path}")
    return matched